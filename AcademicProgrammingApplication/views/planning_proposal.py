from django.http import HttpResponse
from django.shortcuts import render
import pandas as pd
from datetime import datetime
from AcademicProgrammingApplication.models import PlanningProposal
from django.conf import settings
import os
from openpyxl import load_workbook


def planning_proposal(request):
    user = request.user
    file_selected = None

    if request.method == 'POST' and request.FILES.get('file'):
        updated_file = request.FILES['file']
        new_name = f"Programacion_{datetime.now().strftime('%d-%m-%Y_%H%M%S')}.xlsx"
        updated_file.name = new_name

        PlanningProposal.objects.create(username=user.username, name_file=new_name, path=updated_file)

        df = pd.read_excel(updated_file)
        df = df[df['Comentario'].notna()]
        df['Usuario'] = user.username
        df = df[['Nombre_Profesor', 'Fecha_Inicio', 'Comentario', 'Nombre_Materia','Usuario_que_notifica']]
        df['id'] = range(1, len(df) + 1)
        file_selected = df.to_dict(orient='records')

    else:
        file_instance = PlanningProposal.objects.last()
        if file_instance:
            full_file_path = os.path.join(settings.MEDIA_ROOT, str(file_instance.path))
            file_path_with_backslashes = full_file_path.replace('\\', '/')
            workbook = load_workbook(filename=file_path_with_backslashes)
            sheet = workbook.active

            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(row)

            df = pd.read_excel(full_file_path)
            df = df[df['Comentario'].notna()]
            df['Usuario'] = user.username
            df = df[['Nombre_Profesor', 'Fecha_Inicio', 'Comentario', 'Nombre_Materia','Usuario_que_notifica']]
            df['id'] = range(1, len(df) + 1)
            file_selected = df.to_dict(orient='records')


    search_query = request.GET.get('search_query')


    if search_query:
        files = PlanningProposal.objects.filter(username=search_query).order_by('-id')
        if file_selected:
            df = pd.DataFrame(file_selected)
            df = df[df['Usuario_que_notifica'] == search_query] 
            file_selected = df.to_dict(orient='records')
    else:
        files = PlanningProposal.objects.all()


    if request.method == 'GET' and request.GET.get('action') == 'download':
        file_instance = PlanningProposal.objects.last()
        if file_instance:
            full_file_path = os.path.join(settings.MEDIA_ROOT, str(file_instance.path))
            file_path_with_backslashes = full_file_path.replace('\\', '/')

            if os.path.exists(full_file_path):
                with open(full_file_path, 'rb') as file:
                    file_content = file.read()

                response = HttpResponse(file_content, content_type='application/octet-stream')
                response['Content-Disposition'] = 'attachment; filename="%s"' % os.path.basename(full_file_path)
                return response

    return render(request, 'academic-programming-proposal.html', {
        'user_name': user.username,
        'title': 'Propuesta Programacion Academica',
        'files': files,
        'file_selected': file_selected,
        'search_query': search_query,
    })
